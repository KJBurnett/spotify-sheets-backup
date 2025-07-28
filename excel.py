# excel.py

import openpyxl
from openpyxl import load_workbook
import os

EXCEL_FILE_NAME = "song-list.xlsx"

def get_google_sheet_data(sheet):
    """Fetches all data from a Google Sheet worksheet."""
    return sheet.get_all_records()

def reset_excel_from_google_sheet(sheet):
    """
    Overwrites the local Excel file with data from the Google Sheet.
    """
    print(f"Resetting {EXCEL_FILE_NAME} from Google Sheet...")
    
    try:
        # Fetch all values from the Google Sheet to avoid header parsing issues
        all_values = sheet.get_all_values()
        
        if len(all_values) < 4:
            print("Google Sheet has fewer than 4 rows. Cannot determine headers.")
            return

        headers = all_values[3]  # Headers on row 4
        data_rows = all_values[4:]  # Data starts on row 5

        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Music Saved Tracks"

        # Write headers
        ws.append(headers)

        # Write data rows
        for row in data_rows:
            # Pad row if it's shorter than the header length
            if len(row) < len(headers):
                row.extend([''] * (len(headers) - len(row)))
            ws.append(row)

        # Save the workbook
        wb.save(EXCEL_FILE_NAME)
        print(f"Successfully reset {EXCEL_FILE_NAME}.")
        if not data_rows:
            print("Note: The Google Sheet has no data below the header row. The Excel file was created with only headers.")

    except Exception as e:
        print(f"An error occurred during Excel reset: {e}")


def append_row_to_excel(row_data):
    """
    Appends a single row to the local Excel file.
    If the file doesn't exist, it creates it with headers.
    """
    try:
        if not os.path.exists(EXCEL_FILE_NAME):
            print(f"{EXCEL_FILE_NAME} not found. Please use --reset-excel first to create it.")
            return

        wb = load_workbook(EXCEL_FILE_NAME)
        ws = wb.active
        ws.append(row_data)
        wb.save(EXCEL_FILE_NAME)
        print(f"Appended new row to {EXCEL_FILE_NAME}")

    except Exception as e:
        print(f"An error occurred while appending a row to Excel: {e}")

def update_excel_row(song_title, artist, updates):
    """
    Updates specific cells for a row identified by song title and artist.
    `updates` should be a dictionary like {'Acquirement Status': 'acquired', 'Triaged': 'triaged'}.
    """
    try:
        if not os.path.exists(EXCEL_FILE_NAME):
            print(f"{EXCEL_FILE_NAME} not found. Cannot update.")
            return

        wb = load_workbook(EXCEL_FILE_NAME)
        ws = wb.active

        # Find header indices
        headers = [cell.value for cell in ws[1]]
        try:
            title_col_index = headers.index("Title") + 1
            artist_col_index = headers.index("Artist") + 1
            update_indices = {header: headers.index(header) + 1 for header in updates.keys()}
        except ValueError as e:
            print(f"Error: A required column is missing in the Excel file - {e}")
            return

        # Find the row to update
        row_to_update = None
        for row in ws.iter_rows(min_row=2): # Skip header row
            if row[title_col_index - 1].value == song_title and row[artist_col_index - 1].value == artist:
                row_to_update = row
                break
        
        if row_to_update:
            for header, col_index in update_indices.items():
                row_to_update[col_index - 1].value = updates[header]
            wb.save(EXCEL_FILE_NAME)
            print(f"Updated row for '{song_title}' in {EXCEL_FILE_NAME}.")
        else:
            print(f"Could not find row for '{song_title}' in {EXCEL_FILE_NAME} to update.")

    except Exception as e:
        print(f"An error occurred while updating a row in Excel: {e}")
